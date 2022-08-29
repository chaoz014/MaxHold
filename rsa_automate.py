import numpy as np
import matplotlib.pyplot as plt

from rsa import spectrum_example, peak_power_detector


from myFunctions import exportToCSV, readlines, diffList, calc_list_avg, calc_area, calc_max_min
from M825functions import initM825, powerAmplitudedBm, powerAmplitudeQuery, outputON, outputOff, outputQuery, freqSetCWfreq, freqQueryCWstate, pulsePeriodSet, pulseWidthSet, pulseOn, pulsePeriodQuery, pulseWidthQuery, pulseOff
import time

# To read the config file
import pandas as pd
from openpyxl import load_workbook

def stringBool(mystring):
    stringBool = False
    mystring = mystring.upper() 
    if mystring.upper() == "TRUE" :
        stringBool = True
    elif mystring.upper() == "FALSE" :
        stringBool = False
    return stringBool
        
import json
f_json = open("C:\\Users\\berrios1\\Documents\\ARMR\\Tektronix\\RSA_automate\\config.json",)
config_json = json.load(f_json)
f_json.close()
plot_on = stringBool(config_json["SPECTRUM"][0]["plot_on"])
wait_to_continue = stringBool(config_json["SPECTRUM"][0]["wait_to_continue"])
time_to_wait = int(config_json["SPECTRUM"][0]["time_to_wait"])
save_data = stringBool(config_json["DATA"][0]["save_data"])
save_mod_off_once = stringBool(config_json["DATA"][0]["save_mod_off_once"])


if save_data == True:
    print("\nSaving data on.")
else:
    print("\nSaving data off.")

if plot_on == True:
    print("Plot spectrum on.")
else:
    print("Plot spectrum off.")

if wait_to_continue == True :
    print("The program will wait for user to close the plot.\n")
else:
    print("The program will not wait for user to close the plot.\n")

time.sleep(5)

def sweep_n_calc(cf, span, DUT_RBW) :
    trace_points, freq_array, peakPower, peakFreq = spectrum_example(cf, span, DUT_RBW)

    # Calculate area under the curve
    areadB = calc_area(trace_points, freq_array)

    return areadB, peakPower, peakFreq, freq_array, trace_points  

# area, peak_val, peak_freq, freqList, pointList  = sweep_n_calc(2.4453e9, 500e3)

excel_filename = 'C:\\Users\\berrios1\\Documents\\ARMR\\Tektronix\\RSA_automate\\configfile.xlsx'
excel_data_df  = pd.read_excel(excel_filename, sheet_name='config', engine='openpyxl')  # Read in the configuration file

# DUT Configuations
DUT_sweepSpanList = excel_data_df["DUT Span (Hz)"].tolist()
DUT_sweepCenterFreqList = excel_data_df["DUT Center Freq (Hz)"].tolist()
# DUT_traceTypeList = excel_data_df["DUT Trace Type ('string')"].tolist()
DUT_traceTypeList = ["AVERage", "MAXHold", "MINHold", "WRITe"]
DUT_outPutfileNameList = excel_data_df["DUT output filename ('string')"].tolist()
DUT_RBW = 313


print(len(DUT_sweepSpanList), len(DUT_sweepCenterFreqList), len(DUT_traceTypeList), len(DUT_outPutfileNameList))
time.sleep(1)


DUT_traceSet = set(DUT_traceTypeList)
DUT_numTraceCases = len(DUT_traceSet)
print("Number of different DUT traces:", DUT_numTraceCases)
DUT_averageCount = int(excel_data_df["Average Count"].tolist()[0])   # Get the average count for Trace Averaging
print("Average count:", DUT_averageCount)

# check to make sure the DUT has the same number of parameters for each test
if len(DUT_sweepCenterFreqList) == len(DUT_sweepSpanList) and len(DUT_outPutfileNameList) == len(DUT_sweepCenterFreqList) :
    print("DUT test cases match")
else:
    print("ERROR DUT test cases do not match")
    print("len(DUT_sweepCenterFreqList)",len(DUT_sweepCenterFreqList))
    print("len(DUT_sweepSpanList)",len(DUT_sweepSpanList))
    print("len(DUT_outPutfileNameList)",len(DUT_outPutfileNameList))    
    print("len(DUT_traceTypeList)",len(DUT_traceTypeList))   
    quit()

# M825 RF Generator Configurations
M825_inst = initM825()                   # Connect and initialize M825
M825_centerFreqList = excel_data_df['M825 Center Freq (Hz)'].tolist()
M825_pulsePeriodList = excel_data_df['M825 Pulse Period (s)'].tolist()
M825_pulseWidthList = excel_data_df["M825 pulse Width (s)"].tolist()
numTestSets = len(M825_centerFreqList)
M825_modulation = bool(excel_data_df["Modulation"].tolist()[0]) # Get modulation indicator
pulseOff(M825_inst) # M825 - Ensure Pulse is Off 

# check to make sure the M825 has the same number of parameters for each test
if len(M825_centerFreqList) == len(M825_pulsePeriodList) and len(M825_pulseWidthList) == len(M825_pulsePeriodList) :
    print("M825 test cases match")
else:
    print("ERROR M825 test cases do not match")
    print("len(M825_centerFreqList)",len(M825_centerFreqList))
    print("len(M825_pulsePeriodList)",len(M825_pulsePeriodList))
    print("len(M825_pulseWidthList)",len(M825_pulseWidthList))   
    quit()

# Initialize lists and variables first used withint the inner for loop
areaListOnAvg = []
areaListOffAvg = []
peak_valListOnAvg = []
peak_valListOffAvg = []
peak_freqListOnAvg = []
peak_freqListOffAvg = []
areaListOnMax = []
areaListOffMax = []
peak_valListOnMax = []
peak_valListOffMax = []
peak_freqListOnMax = []
peak_freqListOffMax = []
areaListOnMin = []
areaListOffMin = []
peak_valListOnMin = []
peak_valListOffMin = []
peak_freqListOnMin = []
peak_freqListOffMin = []

areaOnWriteCalcAvgList = []
areaOffWriteCalcAvgList = []
peak_valListOnWriteCalcWriteAvg = []
peak_valListOffWriteCalcWriteAvg = []
peak_freqListOnCalcAvg = []
peak_freqListOffCalcAvg = []


areaListOnWrite        = []
peak_valListOnWrite    = []
peak_freqListOnWrite   = []

areaListOffWrite        = []
peak_valListOffWrite    = []
peak_freqListOffWrite   = []

area = ''
peak_val = ''
peak_freq = ''
freqList = []
pointList = []

headerTuple = {("freq (Hz)","Mag (dB)")}

powerAmplitudedBm(M825_inst, -10.0)     # Set the power amplitude (dBm)
print(powerAmplitudeQuery(M825_inst))
outputON(M825_inst)                     # M825 - Turn the output on
print(outputQuery(M825_inst))           # M825 - Query the output
print("Number of test sets",numTestSets)

time.sleep(5)
for i in range(numTestSets): # i is for the M825
    print("i:",i)

    # Set up M825 RF Generator's frequency, pulse period, and pulse width
    print("M825: Next center frequency:", float(M825_centerFreqList[i]))    # M825 - Set the center frequency
    freqSetCWfreq(M825_inst, float(M825_centerFreqList[i]))                 # M825 - Set center frequency
    print(freqQueryCWstate(M825_inst))                                      # M825 - Query the CW frequency

    dut_center_freq = float(DUT_sweepCenterFreqList[i])
    dut_sweep_span = float(DUT_sweepSpanList[i])

    print("\nDUT Center Freq [Hz]:",type(dut_center_freq),dut_center_freq)
    print("DUT Span [Hz]:",type(dut_sweep_span), dut_sweep_span,"\n")
    if pd.isna(dut_center_freq) == True or pd.isna(dut_center_freq) == True :
        print("\nERROR: Clean the end of the excel config sheet. Got a nan.\n")
        break

    # for l in range(DUT_numTraceCases) :
    if M825_modulation == True:
        pulsePeriodSet(M825_inst, float(M825_pulsePeriodList[i]))           # M825 - Set the Pulse Period
        print(pulsePeriodQuery(M825_inst))                                  # M825 - Query Pulse Period
        pulseWidthSet(M825_inst, float(M825_pulseWidthList[i]))             # M825 - Set Pulse Width
        print(pulseWidthQuery(M825_inst))                                   # M825 - Query Pulse Width
        
        # Pulse On
        print("Modulation On")
        pulseOn(M825_inst)  # M825 - Turn Pulse on      

        # Intialize temp lists to store values that will be used to calculate max, min, average ect.
        areaTemp = []
        peakTemp = []
        peakFreqTemp = []
        pointListTemp = []

        #peak_power_detector(freq, trace)
        for k in range(DUT_averageCount):
            print("Mod ON")
            area, peak_val, peak_freq, freqList, pointList  = sweep_n_calc(dut_center_freq, dut_sweep_span, DUT_RBW) # DUT - Initiate sweep

            areaTemp.append(area)
            peakTemp.append(peak_val)
            peakFreqTemp.append(peak_freq)
            pointListTemp.append(pointList)
            time.sleep(.07)

        pointListMax, pointListMin = calc_max_min(pointListTemp)
        peakPowerMax, peakFreqMax = peak_power_detector(freqList, pointListMax)
        peakPowerMin, peakFreqMin = peak_power_detector(freqList, pointListMin)

        pointListAvg = calc_list_avg(pointListTemp, DUT_averageCount)
        areadB_avg = calc_area(pointListAvg, freqList)
        peakPowerAvg, peakFreqAvg = peak_power_detector(freqList, pointListAvg)
        
        if plot_on == True:
            """################PLOT################"""
            # Plot out the four spectrum traces
            fig = plt.figure(1, figsize=(15, 10))
            ax1 = plt.subplot(111)
            ax1.set_title('Spectrum Traces')
            ax1.set_xlabel('Frequency (GHz)')
            ax1.set_ylabel('Amplitude (dBm)')
            freqList /= 1e9
            st1, = plt.plot(freqList, pointListMax)
            st2, = plt.plot(freqList, pointListMin)
            st3, = plt.plot(freqList, pointListAvg)
            st4, = plt.plot(freqList, pointList)
            ax1.legend([st1, st2, st3, st4], ['Max Hold', 'Min Hold', 'Average' ,'Write'])
            ax1.set_xlim([freqList[0], freqList[-1]])  
            ax1.set_ylim([-100,0])
            
            if wait_to_continue == True:
                plt.show(block=False)
                plt.pause(time_to_wait)
                plt.close()
            else:
                plt.show(block=False)
                plt.pause(1)
                plt.close()


        # Store the values depending on trace type
        areaListOnAvg.append(0)
        peak_valListOnAvg.append(0)
        peak_freqListOnAvg.append(0)

        areaListOnMax.append(calc_area(pointListMax, freqList))
        peak_valListOnMax.append(peakPowerMax)
        peak_freqListOnMax.append(peakFreqMax)

        areaListOnMin.append(calc_area(pointListMin, freqList))
        peak_valListOnMin.append(peakPowerMin)
        peak_freqListOnMin.append(peakFreqMin) 

        areaListOnWrite.append(area)
        peak_valListOnWrite.append(peak_val)
        peak_freqListOnWrite.append(peak_freq)

        areaOnWriteCalcAvgList.append(calc_area(pointListAvg, freqList))
        peak_valListOnWriteCalcWriteAvg.append(peakPowerAvg) 
        peak_freqListOnCalcAvg.append(peakFreqAvg)
        
        dataOutFileName = str(DUT_outPutfileNameList[i]) + "_iteration-" + str(i) +  "_Mod-ON.txt" 
        try:
            # Write frequency and magnitude list to file
            dataTuple = zip(freqList, pointList)                        
            exportToCSV(dataTuple, dataOutFileName, headerTuple)         
        except:
            print("Failed try statement: exportToCSV(dataTuple, dataOutFileName, headerTuple)")            
        
        pulseOff(M825_inst) # M825 - Turn Pulse Off            
            
    
    # Pulse Off
    print("Modulation Off")
    time.sleep(2)

    # Intialize temp lists to store values that will be used to calculate max, min, average ect.
    areaTemp = []
    peakTemp = []
    peakFreqTemp = []
    pointListTemp = []

    #peak_power_detector(freq, trace)
    for k in range(DUT_averageCount):
        print("Mod OFF")
        area, peak_val, peak_freq, freqList, pointList  = sweep_n_calc(dut_center_freq, dut_sweep_span, DUT_RBW) # DUT - Initiate sweep

        areaTemp.append(area)
        peakTemp.append(peak_val)
        peakFreqTemp.append(peak_freq)
        pointListTemp.append(pointList)
        time.sleep(.07)


    pointListMax, pointListMin = calc_max_min(pointListTemp)
    peakPowerMax, peakFreqMax = peak_power_detector(freqList, pointListMax)
    peakPowerMin, peakFreqMin = peak_power_detector(freqList, pointListMin)

    pointListAvg = calc_list_avg(pointListTemp, DUT_averageCount)
    areadB_avg = calc_area(pointListAvg, freqList)
    peakPowerAvg, peakFreqAvg = peak_power_detector(freqList, pointListAvg)
    
    if plot_on == True:
        """################PLOT################"""
        # Plot out the four spectrum traces
        fig = plt.figure(1, figsize=(15, 10))
        ax1 = plt.subplot(111)
        ax1.set_title('Spectrum Traces')
        ax1.set_xlabel('Frequency (GHz)')
        ax1.set_ylabel('Amplitude (dBm)')
        freqList /= 1e9
        st1, = plt.plot(freqList, pointListMax)
        st2, = plt.plot(freqList, pointListMin)
        st3, = plt.plot(freqList, pointListAvg)
        st4, = plt.plot(freqList, pointList)
        ax1.legend([st1, st2, st3, st4], ['Max Hold', 'Min Hold', 'Average' ,'Write'])
        ax1.set_xlim([freqList[0], freqList[-1]])  
        ax1.set_ylim([-100,0])
        if wait_to_continue == True:
            plt.show(block=False)
            plt.pause(time_to_wait)
            plt.close()
        else:
            plt.show(block=False)
            plt.pause(1)
            plt.close()


    # Store the values depending on trace type
    areaListOffAvg.append(0)
    peak_valListOffAvg.append(0)
    peak_freqListOffAvg.append(0)

    areaListOffMax.append(calc_area(pointListMax, freqList))
    peak_valListOffMax.append(peakPowerMax)
    peak_freqListOffMax.append(peakFreqMax)

    areaListOffMin.append(calc_area(pointListMin, freqList))
    peak_valListOffMin.append(peakPowerMin)
    peak_freqListOffMin.append(peakFreqMin)   

    areaListOffWrite.append(area)
    peak_valListOffWrite.append(peak_val)
    peak_freqListOffWrite.append(peak_freq)  

    areaOffWriteCalcAvgList.append(calc_area(pointListAvg, freqList))
    peak_valListOffWriteCalcWriteAvg.append(peakPowerAvg)
    peak_freqListOffCalcAvg.append(peakFreqAvg)            
    
    dataOutFileName = str(DUT_outPutfileNameList[i]) + "_iteration-" + str(i) +  "_Mod-OFF.txt"
    try:
        # Write frequency and magnitude list to file
        dataTuple = zip(freqList, pointList)          
        exportToCSV(dataTuple, dataOutFileName, headerTuple)
    except:
        print(type(dataOutFileName),dataOutFileName)
        print("Failed try statement: exportToCSV(dataTuple, dataOutFileName, headerTuple)")
            
if save_data == True:            
    deltaAreaAvgList = diffList(areaListOnAvg,areaListOffAvg)
    deltaPeakAvgList = diffList(peak_valListOnAvg,peak_valListOffAvg)

    deltaAreaMaxList = diffList(areaListOnMax,areaListOffMax)
    deltaPeakMaxList = diffList(peak_valListOnMax,peak_valListOffMax)

    deltaAreaMinList = diffList(areaListOnMin,areaListOffMin)
    deltaPeakMinList = diffList(peak_valListOnMin,peak_valListOffMin)

    deltaAreaWriteList = diffList(areaListOnWrite,areaListOffWrite)
    deltaPeakWriteList = diffList(peak_valListOnWrite,peak_valListOffWrite)

    deltaAreaAvgListCalc = diffList(areaOnWriteCalcAvgList,areaOffWriteCalcAvgList)
    deltaPeakAvgListCalc = diffList(peak_valListOnWriteCalcWriteAvg,peak_valListOffWriteCalcWriteAvg)

    if M825_modulation == True:
        print(len(areaListOnAvg),len(areaListOffAvg),len(peak_valListOnAvg),len(peak_valListOffAvg),len(peak_freqListOnAvg), len(peak_freqListOffAvg),len(areaListOnMax),len(areaListOffMax), len(peak_valListOnMax),
            len(peak_valListOffMax),len(peak_freqListOnMax),len(peak_freqListOffMax),len(areaListOnMin),len(areaListOffMin), len(peak_valListOnMin),len(peak_valListOffMin),len(peak_freqListOnMin), len(peak_freqListOffMin),
            len(areaListOnWrite),len(areaListOffWrite),len(peak_valListOnWrite),len(peak_valListOffWrite),len(peak_freqListOnWrite), len(peak_freqListOffWrite))
        df_all_data = pd.DataFrame(
                            {"areadB On Avg"            :   areaListOnAvg,                        # 1
                            "areadB Off Avg"           :   areaListOffAvg,                       # 2
                            "peak On (dBm) Avg"        :   peak_valListOnAvg,                    # 3
                            "peak Off (dBm) Avg"       :   peak_valListOffAvg,                   # 4
                            "peakFreq On (MHz) Avg"    :   peak_freqListOnAvg,                   # 5
                            "peakFreq Off (MHz) Avg"   :   peak_freqListOffAvg,                  # 6
                            "areadB On Max"            :   areaListOnMax,                        # 7
                            "areadB Off Max"           :   areaListOffMax,                       # 8
                            "peak On (dBm) Max"        :   peak_valListOnMax,                    # 9
                            "peak Off (dBm) Max"       :   peak_valListOffMax,                   # 10
                            "peakFreq On (MHz) Max"    :   peak_freqListOnMax,                   # 11
                            "peakFreq Off (MHz) Max"   :   peak_freqListOffMax,                  # 12
                            "areadB On Min"            :   areaListOnMin,                        # 13
                            "areadB Off Min"           :   areaListOffMin,                       # 14
                            "peak On (dBm) Min"        :   peak_valListOnMin,                    # 15
                            "peak Off (dBm) Min"       :   peak_valListOffMin,                   # 16
                            "peakFreq On (MHz) Min"    :   peak_freqListOnMin,                   # 17
                            "peakFreq Off (MHz) Min"   :   peak_freqListOffMin,                  # 18
                            "areadB On Write"          :   areaListOnWrite,                      # 19
                            "areadB Off Write"         :   areaListOffWrite,                     # 20
                            "peak On (dBm) Write"      :   peak_valListOnWrite,                  # 21
                            "peak Off (dBm) Write"     :   peak_valListOffWrite,                 # 22
                            "peakFreq On (MHz) Write"  :   peak_freqListOnWrite,                 # 23
                            "peakFreq Off (MHz) Write" :   peak_freqListOffWrite,                # 24
                            "Delta Area Avg"           :   deltaAreaAvgList,                     # 25
                            "Delta Area Max"           :   deltaAreaMaxList,                     # 26
                            "Delta Area Min"           :   deltaAreaMinList,                     # 27
                            "Delta Area Write"         :   deltaAreaWriteList,                   # 28
                            "Delta Area Calc Avg"      :   deltaAreaAvgListCalc,                 # 29
                            "Delta Peak Power Avg"     :   deltaPeakAvgList,                     # 30
                            "Delta Peak Power Max"     :   deltaPeakMaxList,                     # 31
                            "Delta Peak Power Min"     :   deltaPeakMinList,                     # 32
                            "Delta Peak Power Write"   :   deltaPeakWriteList,                   # 33
                            "Delta Peak Power Avg Calc":   deltaPeakAvgListCalc,                 # 34
                            "areadB On Calc Avg"       :   areaOnWriteCalcAvgList,               # 35
                            "areadB Off Calc Avg"      :   areaOffWriteCalcAvgList,              # 36
                            "peakFreq On (MHz) Calc Avg" : peak_freqListOnCalcAvg,               # 37
                            "peakFreq Off (MHz) Calc Avg": peak_freqListOffCalcAvg,              # 38
                            "peak On (dBm) Calc Avg"   :   peak_valListOnWriteCalcWriteAvg,      # 39
                            "peak Off (dBm) Calc Avg"  :   peak_valListOffWriteCalcWriteAvg},    # 40
                            
                            columns=["areadB On Avg",                # 1
                                        "areadB Off Avg",               # 2
                                        "peak On (dBm) Avg",            # 3
                                        "peak Off (dBm) Avg",           # 4
                                        "peakFreq On (MHz) Avg",        # 5
                                        "peakFreq Off (MHz) Avg",       # 6
                                        "areadB On Max",                # 7
                                        "areadB Off Max",               # 8
                                        "peak On (dBm) Max",            # 9
                                        "peak Off (dBm) Max",           # 10
                                        "peakFreq On (MHz) Max",        # 11
                                        "peakFreq Off (MHz) Max",       # 12
                                        "areadB On Min",                # 13
                                        "areadB Off Min",               # 14
                                        "peak On (dBm) Min",            # 15
                                        "peak Off (dBm) Min",           # 16
                                        "peakFreq On (MHz) Min",        # 17
                                        "peakFreq Off (MHz) Min",       # 18
                                        "areadB On Write",              # 19
                                        "areadB Off Write",             # 20
                                        "peak On (dBm) Write",          # 21
                                        "peak Off (dBm) Write",         # 22
                                        "peakFreq On (MHz) Write",      # 23
                                        "peakFreq Off (MHz) Write",     # 24
                                        "Delta Area Avg",               # 25
                                        "Delta Area Max",               # 26
                                        "Delta Area Min",               # 27
                                        "Delta Area Write",             # 28
                                        "Delta Area Calc Avg",          # 29
                                        "Delta Peak Power Avg",         # 30
                                        "Delta Peak Power Max",         # 31
                                        "Delta Peak Power Min",         # 32
                                        "Delta Peak Power Write",       # 33
                                        "Delta Peak Power Avg Calc",    # 34
                                        "areadB On Calc Avg",           # 35
                                        "areadB Off Calc Avg",          # 36
                                        "peakFreq On (MHz) Calc Avg",   # 37
                                        "peakFreq Off (MHz) Calc Avg",  # 38
                                        "peak On (dBm) Calc Avg",       # 39
                                        "peak Off (dBm) Calc Avg"])     # 40
    else:
        print(len(areaListOffAvg),len(peak_valListOffAvg),len(peak_freqListOffAvg),len(areaListOffMax),len(peak_valListOffMax), len(peak_freqListOffMax),len(areaListOffMin),len(peak_valListOffMin), len(peak_freqListOffMin))
        df_all_data = pd.DataFrame({
                            "areadB Off Avg"           :   areaListOffAvg,           # 1
                            "peak Off (dBm) Avg"        :   peak_valListOffAvg,      # 2
                            "peakFreq Off (MHz) Avg"   :   peak_freqListOffAvg,      # 3
                            "areadB Off Max"           :   areaListOffMax,           # 4
                            "peak Off (dBm) Max"        :   peak_valListOffMax,      # 5
                            "peakFreq Off (MHz) Max"   :   peak_freqListOffMax,      # 6
                            "areadB Off Min"           :   areaListOffMin,           # 7
                            "peak Off (dBm) Min"        :   peak_valListOffMin,      # 8
                            "peakFreq Off (MHz) Min"   :   peak_freqListOffMin,      # 9
                            "areadB Off Write"         :   areaListOffWrite,         # 10
                            "peak Off (dBm) Write"      :   peak_valListOffWrite,    # 11
                            "peakFreq Off (MHz) Write" :   peak_freqListOffWrite,    # 12
                            "areadB Off Calc Avg"      :   areaOffWriteCalcAvgList}, # 13
                            
                            columns=["areadB Off Avg",           # 1
                                    "peak Off (dBm) Avg",        # 2
                                    "peakFreq Off (MHz) Avg",    # 3
                                    "areadB Off Max",            # 4
                                    "peak Off (dBm) Max",        # 5
                                    "peakFreq Off (MHz) Max",    # 6
                                    "areadB Off Min",            # 7
                                    "peak Off (dBm) Min",        # 8
                                    "peakFreq Off (MHz) Min",    # 9
                                    "areadB Off Write",          # 10
                                    "peak Off (dBm) Write",      # 11
                                    "peakFreq Off (MHz) Write",  # 12
                                    "areadB Off Calc Avg"])      # 13

    workbook = load_workbook(excel_filename)
    # with pd.ExcelWriter(excel_filename, engine='openpyxl', mode='a',if_sheet_exists='new') as writer:
    with pd.ExcelWriter(excel_filename, engine='openpyxl', mode='a') as writer:
        # df_all_data.to_excel(writer,
        #                    sheet_name="data",
        #                    startrow=workbook['data'].max_row,
        #                    header=True,
        #                    index=True)
        df_all_data.to_excel(writer,
                        sheet_name="data")                  

# Write data to file
# dataTuple = zip(areaListOnAvg, areaListOffAvg, peak_valListOnAvg, peak_valListOffAvg, peak_freqListOnAvg, peak_freqListOffAvg, areaListOnMax, areaListOffMax, peak_valListOnMax, peak_valListOffMax, peak_freqListOnMax, peak_freqListOffMax, areaListOnMin, areaListOffMin, peak_valListOnMin, peak_valListOffMin, peak_freqListOnMin, peak_freqListOffMin)
# headerTuple = {("areadB On Avg", "areadB Off Avg", "peak On (Hz) Avg", "peak Off (Hz) Avg", "peakFreq On (MHz) Avg", "peakFreq Off (MHz) Avg","areadB On Max", "areadB Off MAx", "peak On (Hz) Max", "peak Off (Hz) Max", "peakFreq On (MHz) Max", "peakFreq Off (MHz) MAx", "areadB On Min", "areadB Off Min", "peak On (Hz) Min", "peak Off (Hz) Min", "peakFreq On (MHz) Min", "peakFreq Off (MHz) Min")}
# exportToCSV(dataTuple, "DUT_data.txt", headerTuple)       

outputOff(M825_inst)             # M825 - Turn the ouput off
print(outputQuery(M825_inst))    # M825 - Query the output
print("Closing M825 Visa instance..",M825_inst.close())


# Write device setting to file
# Write magnitudes and freqs to file (to save data)


# title = 'peak: {:.1f} dBm at {:.3f} MHz '.format(peak_val, peak_freq/1e6)

# # Convert area to dB
# area = 10*np.log10(area)
# print("Area",area)

# plt.figure(1)
# plt.clf()
# plt.plot(freq_array, points)
# plt.xlabel('freq [MHz]')
# plt.ylabel('dBm')
# plt.grid(True)
# # plt.ylim((min(filter(lambda f: f > 1.0 * min(points), points)), 0))
# plt.ylim(-100, 0.0)

# plt.title(title)
# plt.show()